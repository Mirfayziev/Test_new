"""
AF IMPERIYA - TO'LIQ ROUTES
Barcha 15 ta talab bilan to'liq implementatsiya
"""

from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, send_from_directory
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import json

from modules.models import *
from modules.utils import role_required, module_required, save_file, delete_file


def init_routes(app, db):
    """Barcha route'larni ro'yxatga olish"""
    
    # ==================== 1. TOPSHIRIQLAR (TASKS) ====================
    # Talab #14 va #15 implement qilingan
    
    @app.route('/tasks')
    @login_required
    def tasks():
        """Topshiriqlar ro'yxati - FAQAT O'ZINIKI (Talab #14)"""
        if current_user.role in ['admin', 'rahbar']:
            # Admin va rahbar - barchasi
            tasks = Task.query.filter_by(is_active=True).order_by(Task.created_at.desc()).all()
        else:
            # Xodim - faqat o'ziniki
            tasks = Task.query.filter_by(
                assigned_to_id=current_user.id,
                is_active=True
            ).order_by(Task.created_at.desc()).all()
        
        return render_template('tasks/list.html', tasks=tasks)
    
    @app.route('/tasks/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def task_new():
        """Yangi topshiriq yaratish"""
        if request.method == 'POST':
            task = Task(
                title=request.form.get('title'),
                description=request.form.get('description'),
                priority=request.form.get('priority', 'medium'),
                status='pending',
                created_by_id=current_user.id,
                assigned_to_id=request.form.get('assigned_to_id'),
                deadline=datetime.strptime(request.form.get('deadline'), '%Y-%m-%d') if request.form.get('deadline') else None,
                category=request.form.get('category'),
                estimated_hours=float(request.form.get('estimated_hours', 0))
            )
            
            db.session.add(task)
            db.session.commit()
            
            flash('Topshiriq muvaffaqiyatli yaratildi!', 'success')
            return redirect(url_for('tasks'))
        
        # Xodimlar ro'yxati (barcha faol userlar)
        employees = User.query.filter_by(is_active=True).all()
        return render_template('tasks/form.html', task=None, employees=employees)
    
    @app.route('/tasks/<int:id>')
    @login_required
    def task_detail(id):
        """Topshiriq detali"""
        task = Task.query.get_or_404(id)
        
        # Ruxsat tekshirish
        if current_user.role not in ['admin', 'rahbar']:
            if task.assigned_to_id != current_user.id:
                flash('Ruxsat yo\'q!', 'danger')
                return redirect(url_for('tasks'))
        
        # Comments
        comments = TaskComment.query.filter_by(task_id=id).order_by(TaskComment.created_at.desc()).all()
        
        return render_template('tasks/detail.html', task=task, comments=comments)
    
    @app.route('/tasks/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def task_edit(id):
        """Topshiriqni tahrirlash"""
        task = Task.query.get_or_404(id)
        
        if request.method == 'POST':
            task.title = request.form.get('title')
            task.description = request.form.get('description')
            task.priority = request.form.get('priority')
            task.assigned_to_id = request.form.get('assigned_to_id')
            task.deadline = datetime.strptime(request.form.get('deadline'), '%Y-%m-%d') if request.form.get('deadline') else None
            task.category = request.form.get('category')
            task.estimated_hours = float(request.form.get('estimated_hours', 0))
            task.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('Topshiriq yangilandi!', 'success')
            return redirect(url_for('task_detail', id=id))
        
        employees = User.query.filter_by(is_active=True).all()
        return render_template('tasks/form.html', task=task, employees=employees)
    
    @app.route('/tasks/<int:id>/complete', methods=['POST'])
    @login_required
    def task_complete(id):
        """Topshiriqni bajarish"""
        task = Task.query.get_or_404(id)
        
        if current_user.id != task.assigned_to_id and current_user.role not in ['admin', 'rahbar']:
            flash('Ruxsat yo\'q!', 'danger')
            return redirect(url_for('tasks'))
        
        task.status = 'completed'
        task.progress = 100
        task.completed_at = datetime.utcnow()
        task.actual_hours = float(request.form.get('actual_hours', 0))
        
        db.session.commit()
        
        flash('Topshiriq bajarildi!', 'success')
        return redirect(url_for('tasks'))
    
    @app.route('/tasks/<int:id>/comment', methods=['POST'])
    @login_required
    def task_comment(id):
        """Topshiriqqa izoh qo'shish"""
        task = Task.query.get_or_404(id)
        
        comment = TaskComment(
            task_id=id,
            user_id=current_user.id,
            comment=request.form.get('comment')
        )
        
        db.session.add(comment)
        db.session.commit()
        
        flash('Izoh qo\'shildi!', 'success')
        return redirect(url_for('task_detail', id=id))
    
    # ==================== 2. IJRO MODULI ====================
    # Talab #2: Excel export va Calendar
    
    @app.route('/ijro')
    @login_required
    @role_required('admin', 'rahbar')
    def ijro():
        """Ijro moduli - Xodim o'z topshiriqlarini bajaradi"""
        
        # Xodim faqat o'ziga biriktirilgan topshiriqlarni ko'radi
        if current_user.role in ['admin', 'rahbar']:
            # Admin va Rahbar barcha topshiriqlarni ko'radi
            tasks = Task.query.filter_by(is_active=True).order_by(Task.deadline).all()
        else:
            # Xodim faqat o'z topshiriqlarini ko'radi
            tasks = Task.query.filter_by(
                assigned_to_id=current_user.id,
                is_active=True
            ).order_by(Task.deadline).all()
        
        # Statistika
        stats = {
            'total': len(tasks),
            'pending': len([t for t in tasks if t.status == 'pending']),
            'in_progress': len([t for t in tasks if t.status == 'in_progress']),
            'completed': len([t for t in tasks if t.status == 'completed']),
        }
        
        return render_template('ijro/index.html', tasks=tasks, stats=stats)
    
    @app.route('/ijro/export')
    @login_required
    @role_required('admin', 'rahbar')
    def ijro_export():
        """Ijro Excel export - Talab #2"""
        tasks = Task.query.filter_by(is_active=True).all()
        
        # Workbook yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Ijro"
        
        # Header
        headers = ['№', 'Topshiriq', 'Xodim', 'Boshlangan', 'Muddat', 'Status', 'Progress %', 'Ustunlik']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4A6EF5", end_color="4A6EF5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ma'lumotlar
        for idx, task in enumerate(tasks, 1):
            ws.append([
                idx,
                task.title,
                task.assigned_to.full_name if task.assigned_to else '-',
                task.created_at.strftime('%d.%m.%Y'),
                task.deadline.strftime('%d.%m.%Y') if task.deadline else '-',
                task.status,
                task.progress or 0,
                task.priority
            ])
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name=f'ijro_{datetime.now().strftime("%Y%m%d")}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    @app.route('/api/ijro/calendar')
    @login_required
    @role_required('admin', 'rahbar')
    def ijro_calendar_data():
        """Calendar uchun ma'lumotlar - Talab #2"""
        tasks = Task.query.filter_by(is_active=True).all()
        
        events = []
        for task in tasks:
            if task.deadline:
                color = {
                    'high': '#FF3B30',
                    'medium': '#FF9500',
                    'low': '#00CC99'
                }.get(task.priority, '#4A6EF5')
                
                events.append({
                    'id': task.id,
                    'title': task.title,
                    'start': task.deadline.isoformat(),
                    'color': color,
                    'url': url_for('task_detail', id=task.id)
                })
        
        return jsonify(events)
    
    @app.route('/api/task/<int:id>/detail')
    @login_required
    def api_task_detail(id):
        """Task detail API - Talab #15"""
        task = Task.query.get_or_404(id)
        
        return jsonify({
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'employee': task.assigned_to.full_name if task.assigned_to else '-',
            'progress': task.progress or 0,
            'status': task.status,
            'priority': task.priority,
            'deadline': task.deadline.strftime('%d.%m.%Y') if task.deadline else '-',
            'created_at': task.created_at.strftime('%d.%m.%Y %H:%M'),
            'comments_count': TaskComment.query.filter_by(task_id=id).count()
        })
    
    # ==================== VEHICLES (AVTO TRANSPORT) ====================
    # Talab #3: Rasm va PDF yuklash
    
    @app.route('/vehicles')
    @login_required
    @role_required('admin', 'rahbar')
    def vehicles():
        """Avto transport ro'yxati"""
        vehicles = Vehicle.query.filter_by(is_active=True).all()
        return render_template('vehicles/list.html', vehicles=vehicles)
    
    @app.route('/vehicles/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def vehicle_new():
        """Yangi transport qo'shish - Talab #3"""
        if request.method == 'POST':
            vehicle = Vehicle(
                name=request.form.get('name'),
                brand=request.form.get('brand'),
                model=request.form.get('model'),
                year=int(request.form.get('year')) if request.form.get('year') else None,
                plate_number=request.form.get('plate_number'),
                vin_number=request.form.get('vin_number'),
                color=request.form.get('color'),
                fuel_type=request.form.get('fuel_type'),
                category=request.form.get('category'),
                mileage=int(request.form.get('mileage', 0)),
                status=request.form.get('status', 'active')
            )
            
            # Rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    vehicle.photo_path = save_file(file, 'vehicles', 'vehicle_photo')
            
            # PDF yuklash
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    vehicle.documents_path = save_file(file, 'vehicles', 'vehicle_doc')
            
            db.session.add(vehicle)
            db.session.commit()
            
            flash('Transport qo\'shildi!', 'success')
            return redirect(url_for('vehicles'))
        
        return render_template('vehicles/form.html', vehicle=None)
    
    @app.route('/vehicles/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def vehicle_edit(id):
        """Transportni tahrirlash"""
        vehicle = Vehicle.query.get_or_404(id)
        
        if request.method == 'POST':
            vehicle.name = request.form.get('name')
            vehicle.brand = request.form.get('brand')
            vehicle.model = request.form.get('model')
            vehicle.year = int(request.form.get('year')) if request.form.get('year') else None
            vehicle.plate_number = request.form.get('plate_number')
            vehicle.vin_number = request.form.get('vin_number')
            vehicle.color = request.form.get('color')
            vehicle.fuel_type = request.form.get('fuel_type')
            vehicle.category = request.form.get('category')
            vehicle.mileage = int(request.form.get('mileage', 0))
            vehicle.status = request.form.get('status')
            vehicle.updated_at = datetime.utcnow()
            
            # Yangi rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    # Eski faylni o'chirish
                    if vehicle.photo_path:
                        delete_file(vehicle.photo_path)
                    vehicle.photo_path = save_file(file, 'vehicles', 'vehicle_photo')
            
            # Yangi PDF yuklash
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    if vehicle.documents_path:
                        delete_file(vehicle.documents_path)
                    vehicle.documents_path = save_file(file, 'vehicles', 'vehicle_doc')
            
            db.session.commit()
            
            flash('Transport yangilandi!', 'success')
            return redirect(url_for('vehicles'))
        
        return render_template('vehicles/form.html', vehicle=vehicle)
    
    # ==================== BUILDINGS (BINOLAR) ====================
    # Talab #4: Rasm va PDF yuklash
    
    @app.route('/buildings')
    @login_required
    @role_required('admin', 'rahbar')
    def buildings():
        """Binolar ro'yxati"""
        buildings = Building.query.filter_by(is_active=True).all()
        return render_template('buildings/list.html', buildings=buildings)
    
    @app.route('/buildings/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def building_new():
        """Yangi bino qo'shish - Talab #4"""
        if request.method == 'POST':
            building = Building(
                name=request.form.get('name'),
                address=request.form.get('address'),
                category=request.form.get('category'),
                area=float(request.form.get('area', 0)),
                floors=int(request.form.get('floors', 0)),
                rooms=int(request.form.get('rooms', 0)),
                construction_year=int(request.form.get('construction_year')) if request.form.get('construction_year') else None,
                condition=request.form.get('condition'),
                description=request.form.get('description')
            )
            
            # Rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    building.photo_path = save_file(file, 'buildings', 'building_photo')
            
            # PDF yuklash
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    building.documents_path = save_file(file, 'buildings', 'building_doc')
            
            # Blueprint yuklash
            if 'blueprint' in request.files:
                file = request.files['blueprint']
                if file and file.filename:
                    building.blueprint_path = save_file(file, 'buildings', 'blueprint')
            
            db.session.add(building)
            db.session.commit()
            
            flash('Bino qo\'shildi!', 'success')
            return redirect(url_for('buildings'))
        
        return render_template('buildings/form.html', building=None)
    
    @app.route('/buildings/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def building_edit(id):
        """Binoni tahrirlash"""
        building = Building.query.get_or_404(id)
        
        if request.method == 'POST':
            building.name = request.form.get('name')
            building.address = request.form.get('address')
            building.category = request.form.get('category')
            building.area = float(request.form.get('area', 0))
            building.floors = int(request.form.get('floors', 0))
            building.rooms = int(request.form.get('rooms', 0))
            building.construction_year = int(request.form.get('construction_year')) if request.form.get('construction_year') else None
            building.condition = request.form.get('condition')
            building.description = request.form.get('description')
            building.updated_at = datetime.utcnow()
            
            # File uploads
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    if building.photo_path:
                        delete_file(building.photo_path)
                    building.photo_path = save_file(file, 'buildings', 'building_photo')
            
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    if building.documents_path:
                        delete_file(building.documents_path)
                    building.documents_path = save_file(file, 'buildings', 'building_doc')
            
            if 'blueprint' in request.files:
                file = request.files['blueprint']
                if file and file.filename:
                    if building.blueprint_path:
                        delete_file(building.blueprint_path)
                    building.blueprint_path = save_file(file, 'buildings', 'blueprint')
            
            db.session.commit()
            
            flash('Bino yangilandi!', 'success')
            return redirect(url_for('buildings'))
        
        return render_template('buildings/form.html', building=building)
    
    @app.route('/api/building/<int:id>/info')
    @login_required
    def api_building_info(id):
        """Bino ma'lumotlari API - Talab #5"""
        building = Building.query.get_or_404(id)
        
        # Yashil makonlar
        green_spaces = GreenSpace.query.filter_by(building_id=id, is_active=True).all()
        
        return jsonify({
            'id': building.id,
            'name': building.name,
            'address': building.address,
            'area': building.area,
            'floors': building.floors,
            'rooms': building.rooms,
            'condition': building.condition,
            'green_spaces_count': len(green_spaces),
            'green_spaces': [{'id': gs.id, 'name': gs.name, 'area': gs.area} for gs in green_spaces]
        })
    # ==================== GREEN SPACES (YASHIL MAKONLAR) ====================
    # Talab #5: Binolarni tanlash
    
    @app.route('/green-spaces')
    @login_required
    @role_required('admin', 'rahbar')
    def green_spaces():
        """Yashil makonlar ro'yxati"""
        spaces = GreenSpace.query.filter_by(is_active=True).all()
        return render_template('green_spaces/list.html', spaces=spaces)
    
    @app.route('/green-spaces/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def green_space_new():
        """Yangi yashil makon - Talab #5"""
        if request.method == 'POST':
            space = GreenSpace(
                name=request.form.get('name'),
                building_id=request.form.get('building_id') if request.form.get('building_id') else None,
                category=request.form.get('category'),
                area=float(request.form.get('area', 0)),
                location=request.form.get('location'),
                plant_types=request.form.get('plant_types'),
                tree_count=int(request.form.get('tree_count', 0)),
                shrub_count=int(request.form.get('shrub_count', 0)),
                maintenance_schedule=request.form.get('maintenance_schedule'),
                responsible_person=request.form.get('responsible_person'),
                notes=request.form.get('notes')
            )
            
            db.session.add(space)
            db.session.commit()
            
            flash('Yashil makon qo\'shildi!', 'success')
            return redirect(url_for('green_spaces'))
        
        # Binolar ro'yxati
        buildings = Building.query.filter_by(is_active=True).all()
        return render_template('green_spaces/form.html', space=None, buildings=buildings)
    
    # ==================== SOLAR PANELS (QUYOSH PANELLARI) ====================
    # Talab #6: API integratsiya
    
    @app.route('/solar-panels')
    @login_required
    @role_required('admin', 'rahbar')
    def solar_panels():
        """Quyosh panellari ro'yxati"""
        panels = SolarPanel.query.filter_by(is_active=True).all()
        return render_template('solar_panels/list.html', panels=panels)
    
    @app.route('/solar-panels/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def solar_panel_new():
        """Yangi quyosh paneli - Talab #6"""
        if request.method == 'POST':
            panel = SolarPanel(
                name=request.form.get('name'),
                building_id=request.form.get('building_id'),
                capacity_kw=float(request.form.get('capacity_kw', 0)),
                panel_count=int(request.form.get('panel_count', 0)),
                brand=request.form.get('brand'),
                model=request.form.get('model'),
                installation_date=datetime.strptime(request.form.get('installation_date'), '%Y-%m-%d') if request.form.get('installation_date') else None,
                api_endpoint=request.form.get('api_endpoint'),
                api_key=request.form.get('api_key'),
                efficiency=float(request.form.get('efficiency', 0))
            )
            
            db.session.add(panel)
            db.session.commit()
            
            flash('Quyosh paneli qo\'shildi!', 'success')
            return redirect(url_for('solar_panels'))
        
        buildings = Building.query.filter_by(is_active=True).all()
        return render_template('solar_panels/form.html', panel=None, buildings=buildings)
    
    @app.route('/api/solar/realtime')
    @login_required
    @role_required('admin', 'rahbar')
    def solar_realtime():
        """Real-time quyosh paneli ma'lumotlari - Talab #6"""
        import requests
        
        panels = SolarPanel.query.filter_by(is_active=True).all()
        
        data = {
            'labels': [],
            'current': [],
            'daily': [],
            'total': 0
        }
        
        for panel in panels:
            data['labels'].append(panel.name)
            
            # API integration
            if panel.api_endpoint:
                try:
                    headers = {}
                    if panel.api_key:
                        headers['Authorization'] = f'Bearer {panel.api_key}'
                    
                    response = requests.get(panel.api_endpoint, headers=headers, timeout=5)
                    if response.status_code == 200:
                        api_data = response.json()
                        panel.current_production = api_data.get('current_power', 0)
                        panel.daily_production = api_data.get('daily_energy', 0)
                        panel.last_sync = datetime.utcnow()
                        db.session.commit()
                except:
                    pass
            
            data['current'].append(panel.current_production or 0)
            data['daily'].append(panel.daily_production or 0)
            data['total'] += (panel.current_production or 0)
        
        return jsonify(data)
    
    # ==================== EMPLOYEES (XODIMLAR) ====================
    # Talab #7: PDF/Word yuklash
    
    @app.route('/employees')
    @login_required
    @role_required('admin', 'rahbar')
    def employees():
        """Xodimlar ro'yxati"""
        employees = Employee.query.filter_by(is_active=True).all()
        return render_template('employees/list.html', employees=employees)
    
    @app.route('/employees/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def employee_new():
        """Yangi xodim - Talab #7"""
        if request.method == 'POST':
            employee = Employee(
                full_name=request.form.get('full_name'),
                position=request.form.get('position'),
                department=request.form.get('department'),
                hire_date=datetime.strptime(request.form.get('hire_date'), '%Y-%m-%d') if request.form.get('hire_date') else None,
                birth_date=datetime.strptime(request.form.get('birth_date'), '%Y-%m-%d') if request.form.get('birth_date') else None,
                passport_number=request.form.get('passport_number'),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                address=request.form.get('address'),
                salary=float(request.form.get('salary', 0)),
                status='active'
            )
            
            # Rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    employee.photo_path = save_file(file, 'employees', 'emp_photo')
            
            # Pasport PDF yuklash
            if 'passport_file' in request.files:
                file = request.files['passport_file']
                if file and file.filename:
                    employee.passport_file = save_file(file, 'employees', 'passport')
            
            # Rezume yuklash
            if 'resume_file' in request.files:
                file = request.files['resume_file']
                if file and file.filename:
                    employee.resume_file = save_file(file, 'employees', 'resume')
            
            # Shartnoma yuklash
            if 'contract_file' in request.files:
                file = request.files['contract_file']
                if file and file.filename:
                    employee.contract_file = save_file(file, 'employees', 'contract')
            
            db.session.add(employee)
            db.session.commit()
            
            flash('Xodim qo\'shildi!', 'success')
            return redirect(url_for('employees'))
        
        return render_template('employees/form.html', employee=None)
    
    @app.route('/employees/<int:id>/passport')
    @login_required
    def employee_passport(id):
        """Xodim pasportini ko'rish/yuklab olish"""
        employee = Employee.query.get_or_404(id)
        if not employee.passport_file:
            flash('Pasport fayli yo\'q', 'warning')
            return redirect(url_for('employees'))
        
        return send_from_directory(
            app.config['UPLOAD_FOLDER'],
            employee.passport_file,
            as_attachment=False
        )
    
    @app.route('/employees/<int:id>/resume')
    @login_required
    def employee_resume(id):
        """Xodim rezumesini ko'rish/yuklab olish"""
        employee = Employee.query.get_or_404(id)
        if not employee.resume_file:
            flash('Rezume fayli yo\'q', 'warning')
            return redirect(url_for('employees'))
        
        return send_from_directory(
            app.config['UPLOAD_FOLDER'],
            employee.resume_file,
            as_attachment=False
        )
    
    # ==================== OUTSOURCING (AUTSORSING) ====================
    # Talab #8: To'liq ma'lumotlar
    
    @app.route('/outsourcing')
    @login_required
    @role_required('admin', 'rahbar')
    def outsourcing():
        """Autsorsing ro'yxati"""
        services = Outsourcing.query.filter_by(is_active=True).all()
        return render_template('outsourcing/list.html', services=services)
    
    @app.route('/outsourcing/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def outsourcing_new():
        """Yangi autsorsing xizmati - Talab #8"""
        if request.method == 'POST':
            service = Outsourcing(
                service_name=request.form.get('service_name'),
                contract_number=request.form.get('contract_number'),
                contract_date=datetime.strptime(request.form.get('contract_date'), '%Y-%m-%d'),
                start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d') if request.form.get('start_date') else None,
                end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d') if request.form.get('end_date') else None,
                amount=float(request.form.get('amount', 0)),
                contact_person=request.form.get('contact_person'),
                contact_phone=request.form.get('contact_phone'),
                contact_email=request.form.get('contact_email'),
                company_name=request.form.get('company_name'),
                description=request.form.get('description'),
                status='active'
            )
            
            db.session.add(service)
            db.session.commit()
            
            flash('Autsorsing xizmati qo\'shildi!', 'success')
            return redirect(url_for('outsourcing'))
        
        return render_template('outsourcing/form.html', service=None)
    
    # ==================== ORGANIZATIONS (TASHKILOTLAR) ====================
    # Talab #9: File upload
    
    @app.route('/organizations')
    @login_required
    @role_required('admin', 'rahbar')
    def organizations():
        """Tashkilotlar ro'yxati"""
        orgs = Organization.query.filter_by(is_active=True).all()
        return render_template('organizations/list.html', organizations=orgs)
    
    @app.route('/organizations/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def organization_new():
        """Yangi tashkilot - Talab #9"""
        if request.method == 'POST':
            org = Organization(
                name=request.form.get('name'),
                tin=request.form.get('tin'),
                address=request.form.get('address'),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                website=request.form.get('website'),
                director=request.form.get('director'),
                contact_person=request.form.get('contact_person'),
                category=request.form.get('category'),
                description=request.form.get('description')
            )
            
            # Logo yuklash
            if 'logo' in request.files:
                file = request.files['logo']
                if file and file.filename:
                    org.logo_path = save_file(file, 'organizations', 'logo')
            
            # Hujjatlar yuklash
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    org.documents_path = save_file(file, 'organizations', 'doc')
            
            # Sertifikatlar yuklash
            if 'certificates' in request.files:
                file = request.files['certificates']
                if file and file.filename:
                    org.certificates_path = save_file(file, 'organizations', 'cert')
            
            db.session.add(org)
            db.session.commit()
            
            flash('Tashkilot qo\'shildi!', 'success')
            return redirect(url_for('organizations'))
        
        return render_template('organizations/form.html', organization=None)
    
    # ==================== GUESTS (MEHMONLAR) ====================
    # Talab #10: File upload
    
    @app.route('/guests')
    @login_required
    @role_required('admin', 'rahbar')
    def guests():
        """Mehmonlar ro'yxati"""
        guests = Guest.query.order_by(Guest.visit_date.desc()).all()
        return render_template('guests/list.html', guests=guests)
    
    @app.route('/guests/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def guest_new():
        """Yangi mehmon - Talab #10"""
        if request.method == 'POST':
            guest = Guest(
                full_name=request.form.get('full_name'),
                organization=request.form.get('organization'),
                position=request.form.get('position'),
                visit_date=datetime.strptime(request.form.get('visit_date'), '%Y-%m-%d %H:%M'),
                visit_purpose=request.form.get('visit_purpose'),
                passport_number=request.form.get('passport_number'),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                host_name=request.form.get('host_name'),
                host_department=request.form.get('host_department'),
                notes=request.form.get('notes')
            )
            
            # Rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    guest.photo_path = save_file(file, 'guests', 'guest_photo')
            
            # Hujjatlar yuklash
            if 'documents' in request.files:
                file = request.files['documents']
                if file and file.filename:
                    guest.documents_path = save_file(file, 'guests', 'guest_doc')
            
            db.session.add(guest)
            db.session.commit()
            
            flash('Mehmon qo\'shildi!', 'success')
            return redirect(url_for('guests'))
        
        return render_template('guests/form.html', guest=None)
    
    # ==================== CONGRATULATIONS (TABRIKNOMALAR) ====================
    # Talab #11: Kategoriya va file upload
    
    @app.route('/congratulations')
    @login_required
    @role_required('admin', 'rahbar')
    def congratulations():
        """Tabriknomalar ro'yxati"""
        congrats = Congratulation.query.order_by(Congratulation.date.desc()).all()
        return render_template('congratulations/list.html', congratulations=congrats)
    
    @app.route('/congratulations/new', methods=['GET', 'POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def congratulation_new():
        """Yangi tabriknoma - Talab #11"""
        if request.method == 'POST':
            congrat = Congratulation(
                category=request.form.get('category'),
                occasion=request.form.get('occasion'),
                date=datetime.strptime(request.form.get('date'), '%Y-%m-%d'),
                recipient_name=request.form.get('recipient_name'),
                gift_description=request.form.get('gift_description'),
                amount=float(request.form.get('amount', 0)),
                message=request.form.get('message'),
                sent_by=current_user.full_name
            )
            
            # Rasm yuklash
            if 'photo' in request.files:
                file = request.files['photo']
                if file and file.filename:
                    congrat.photo_path = save_file(file, 'congratulations', 'congrat_photo')
            
            # Sertifikat PDF yuklash
            if 'certificate' in request.files:
                file = request.files['certificate']
                if file and file.filename:
                    congrat.certificate_path = save_file(file, 'congratulations', 'certificate')
            
            db.session.add(congrat)
            db.session.commit()
            
            flash('Tabriknoma qo\'shildi!', 'success')
            return redirect(url_for('congratulations'))
        
        employees = Employee.query.filter_by(is_active=True).all()
        return render_template('congratulations/form.html', congratulation=None, employees=employees)
    
    # ==================== CONTRACTS (SHARTNOMALAR) ====================
    # Talab #12: Excel export va yaratish
    
    @app.route('/contracts')
    @login_required
    def contracts():
        """Shartnomalar ro'yxati - Talab #14: faqat o'ziniki"""
        if current_user.role in ['admin', 'rahbar']:
            contracts = Contract.query.filter_by(is_active=True).all()
        else:
            contracts = Contract.query.filter_by(created_by_id=current_user.id, is_active=True).all()
        
        return render_template('contracts/list.html', contracts=contracts)
    
    @app.route('/contracts/new', methods=['GET', 'POST'])
    @login_required
    def contract_new():
        """Yangi shartnoma - Talab #12: xodim o'zi yaratadi"""
        if request.method == 'POST':
            contract = Contract(
                contract_number=request.form.get('contract_number'),
                created_by_id=current_user.id,
                company_name=request.form.get('company_name'),
                contract_type=request.form.get('contract_type'),
                start_date=datetime.strptime(request.form.get('start_date'), '%Y-%m-%d') if request.form.get('start_date') else None,
                end_date=datetime.strptime(request.form.get('end_date'), '%Y-%m-%d') if request.form.get('end_date') else None,
                amount=float(request.form.get('amount', 0)),
                currency=request.form.get('currency', 'UZS'),
                status='active',
                description=request.form.get('description')
            )
            
            # Shartnoma fayli yuklash
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename:
                    contract.file_path = save_file(file, 'contracts', 'contract')
            
            db.session.add(contract)
            db.session.commit()
            
            flash('Shartnoma qo\'shildi!', 'success')
            return redirect(url_for('contracts'))
        
        return render_template('contracts/form.html', contract=None)
    
    @app.route('/contracts/export')
    @login_required
    @role_required('admin', 'rahbar')
    def contracts_export():
        """Shartnomalar Excel export - Talab #12"""
        contracts = Contract.query.filter_by(is_active=True).all()
        
        # Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Shartnomalar"
        
        # Header
        headers = ['№', 'Raqam', 'Kompaniya', 'Turi', 'Boshlangan', 'Tugash', 'Summa', 'Status', 'Yaratgan']
        ws.append(headers)
        
        # Style
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="4A6EF5", end_color="4A6EF5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for idx, contract in enumerate(contracts, 1):
            ws.append([
                idx,
                contract.contract_number,
                contract.company_name,
                contract.contract_type,
                contract.start_date.strftime('%d.%m.%Y') if contract.start_date else '',
                contract.end_date.strftime('%d.%m.%Y') if contract.end_date else '',
                contract.amount,
                contract.status,
                contract.created_by.full_name if contract.created_by else ''
            ])
        
        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 25
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            download_name=f'shartnomalar_{datetime.now().strftime("%Y%m%d")}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    # ==================== WAREHOUSE (OMBOR) ====================
    # Talab #13: To'liq ma'lumotlar
    
    @app.route('/warehouse')
    @login_required
    def warehouse():
        """Ombor so'rovnomalari"""
        if current_user.role in ['admin', 'rahbar']:
            requests_list = WarehouseRequest.query.filter_by(is_active=True).all()
        else:
            requests_list = WarehouseRequest.query.filter_by(user_id=current_user.id, is_active=True).all()
        
        return render_template('warehouse/list.html', requests=requests_list)
    
    @app.route('/warehouse/new', methods=['GET', 'POST'])
    @login_required
    def warehouse_new():
        """Yangi ombor so'rovnomasi - Talab #13"""
        if request.method == 'POST':
            # Auto-generate request number
            last_req = WarehouseRequest.query.order_by(WarehouseRequest.id.desc()).first()
            req_number = f"WH{datetime.now().strftime('%Y%m%d')}{(last_req.id + 1) if last_req else 1:04d}"
            
            req = WarehouseRequest(
                request_number=req_number,
                user_id=current_user.id,
                department=request.form.get('department'),
                request_date=datetime.strptime(request.form.get('request_date'), '%Y-%m-%d'),
                needed_date=datetime.strptime(request.form.get('needed_date'), '%Y-%m-%d') if request.form.get('needed_date') else None,
                items=request.form.get('items'),
                quantity=request.form.get('quantity'),
                unit=request.form.get('unit'),
                purpose=request.form.get('purpose'),
                total_items=int(request.form.get('total_items', 0)),
                estimated_cost=float(request.form.get('estimated_cost', 0)),
                priority=request.form.get('priority', 'normal'),
                status='pending',
                notes=request.form.get('notes')
            )
            
            db.session.add(req)
            db.session.commit()
            
            flash('So\'rovnoma yuborildi!', 'success')
            return redirect(url_for('warehouse'))
        
        return render_template('warehouse/form.html', request=None)
    
    @app.route('/warehouse/<int:id>/approve', methods=['POST'])
    @login_required
    @role_required('admin', 'rahbar')
    def warehouse_approve(id):
        """So'rovnomani tasdiqlash"""
        req = WarehouseRequest.query.get_or_404(id)
        req.status = 'approved'
        req.approved_by_id = current_user.id
        req.approval_date = datetime.utcnow()
        
        db.session.commit()
        
        flash('So\'rovnoma tasdiqlandi!', 'success')
        return redirect(url_for('warehouse'))
